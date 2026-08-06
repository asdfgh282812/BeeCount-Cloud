"""分類 / 帳戶匯入解析(2026-08 新增)。

跟交易匯入(`parser.py` + `transformer.py`)不同 —— 範本格式是我們自己定義
的固定欄位,不需要「欄位對應(field mapping)」這層 UI/邏輯,直接照表頭
名稱讀取即可。CSV/Excel 讀取沿用 `parser.py`/`routers/import_data/
endpoints.py` 同款的 utf-8-sig/gbk decode + openpyxl 讀取邏輯,但因為欄位
形狀完全不同,獨立成這個小模組,不reach into `parser.py` 的內部函式。
"""
from __future__ import annotations

import csv
import io

from .schema import ImportAccount, ImportCategory
from .schema import ImportError as ImpError

_CATEGORY_HEADERS = ("名稱", "類型", "父分類")
_ACCOUNT_HEADERS = (
    "名稱", "類型", "幣別", "期初餘額", "主帳戶名稱", "信用額度", "帳單日", "繳款日",
)

_KIND_EXPENSE = {"支出", "expense", "消費", "-", "支"}
_KIND_INCOME = {"收入", "income", "收", "+", "入"}

# 帳戶類型:同時接受三語系 UI 顯示標籤(zh-TW/zh-CN/en)跟內部值,盡量寬容。
_ACCOUNT_TYPE_ALIASES: dict[str, str] = {
"现金": "cash", "現金": "cash", "cash": "cash",
    "银行卡": "bank_card", "銀行卡": "bank_card", "银行": "bank_card", "銀行": "bank_card",
    "bank card": "bank_card", "bank_card": "bank_card", "bank": "bank_card",
    "信用卡": "credit_card", "credit card": "credit_card", "credit_card": "credit_card",
    "主账户": "account_group", "主帳戶": "account_group",
    "帳戶群組": "account_group", "帳戶群組": "account_group", "群組": "account_group", "組": "account_group",
    "主账户(合并账单)": "account_group", "主帳戶(合併帳單)": "account_group",
    "主帳戶(群組)": "account_group", "主账户(群组)": "account_group",
    "parent account (merged billing)": "account_group", "account_group": "account_group",
    "group": "account_group",
    "支付宝": "alipay", "支付寶": "alipay", "alipay": "alipay",
    "微信": "wechat", "wechat": "wechat", "wechat pay": "wechat",
    "其他": "other", "other": "other",
    "不动产": "real_estate", "不動產": "real_estate", "real estate": "real_estate",
    "real_estate": "real_estate",
    "车辆": "vehicle", "車輛": "vehicle", "vehicle": "vehicle",
    "投资理财": "investment", "投資理財": "investment", "投资": "investment",
    "投資": "investment", "investment": "investment",
    "保险": "insurance", "保險": "insurance", "insurance": "insurance",
    "公积金/社保": "social_fund", "公積金/社保": "social_fund",
    "公积金": "social_fund", "公積金": "social_fund",
    "social fund": "social_fund", "social_fund": "social_fund",
    "贷款": "loan", "貸款": "loan", "loan": "loan",
}


def _read_rows_2d(payload: bytes, filename: str) -> list[list[str]]:
    filename_lower = (filename or "").lower()
    is_xlsx = filename_lower.endswith(".xlsx") or payload[:4] == b"PK\x03\x04"
    if is_xlsx:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
        if not wb.sheetnames:
            return []
        ws = wb[wb.sheetnames[0]]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells: list[str] = []
            for v in row:
                if v is None:
                    cells.append("")
                elif isinstance(v, (int, float)):
                    cells.append(str(int(v)) if float(v) == int(v) else str(v))
                else:
                    cells.append(str(v))
            rows.append(cells)
        wb.close()
        return rows

    # 多重文字編碼嘗試機制 (UTF-8-SIG -> UTF-8 -> CP950/Big5 -> GB18030/GBK -> Replace)
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp950", "big5", "gb18030", "gbk"):
        try:
            text = payload.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    if text is None:
        text = payload.decode("utf-8", errors="replace")

    text = text.replace("\r\n", "\n").replace("\r", "\n")
    if not text.strip():
        return []
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


def _header_index_map(headers: list[str], expected: tuple[str, ...]) -> dict[str, int]:
    normalized = {h.strip(): i for i, h in enumerate(headers)}
    return {name: normalized[name] for name in expected if name in normalized}


def _cell(raw: list[str], col: dict[str, int], key: str) -> str:
    idx = col.get(key)
    if idx is None or idx >= len(raw):
        return ""
    return (raw[idx] or "").strip()


def parse_categories_file(
    payload: bytes, filename: str,
) -> tuple[list[ImportCategory], list[ImpError]]:
    rows_2d = _read_rows_2d(payload, filename)
    if not rows_2d:
        return [], [ImpError(code="IMPORT_NO_ROWS", row_number=0, message="empty file")]
    headers = [str(h).strip() for h in rows_2d[0]]
    col = _header_index_map(headers, _CATEGORY_HEADERS)
    if "名稱" not in col or "類型" not in col:
        return [], [
            ImpError(
                code="IMPORT_TEMPLATE_MISMATCH", row_number=1,
                message="找不到「名稱」或「類型」欄位，請使用最新範本",
            )
        ]

    out: list[ImportCategory] = []
    errors: list[ImpError] = []
    seen: set[tuple[str, str]] = set()
    for offset, raw in enumerate(rows_2d[1:], start=2):
        if not raw or all((c or "").strip() == "" for c in raw):
            continue
        name = _cell(raw, col, "名稱")
        kind_raw = _cell(raw, col, "類型").lower()
        parent = _cell(raw, col, "父分類")
        if not name:
            errors.append(ImpError(
                code="IMPORT_MISSING_NAME", row_number=offset, message="名稱為必填", field_name="name",
            ))
            continue
        if kind_raw in _KIND_EXPENSE:
            kind = "expense"
        elif kind_raw in _KIND_INCOME:
            kind = "income"
        else:
            errors.append(ImpError(
                code="IMPORT_INVALID_KIND", row_number=offset,
                message=f"類型「{kind_raw}」無法辨識，請填「支出」或「收入」", field_name="kind",
            ))
            continue
        key = (name.lower(), kind)
        if key in seen:
            errors.append(ImpError(
                code="IMPORT_DUPLICATE_ROW", row_number=offset,
                message=f"分類「{name}」在檔案中重複", field_name="name",
            ))
            continue
        seen.add(key)
        out.append(ImportCategory(
            name=name, kind=kind, parent_name=parent or None, source_row_number=offset,
        ))
    return out, errors


def parse_accounts_file(
    payload: bytes, filename: str,
) -> tuple[list[ImportAccount], list[ImpError]]:
    rows_2d = _read_rows_2d(payload, filename)
    if not rows_2d:
        return [], [ImpError(code="IMPORT_NO_ROWS", row_number=0, message="empty file")]
    headers = [str(h).strip() for h in rows_2d[0]]
    col = _header_index_map(headers, _ACCOUNT_HEADERS)
    if "名稱" not in col or "類型" not in col or "幣別" not in col:
        return [], [
            ImpError(
                code="IMPORT_TEMPLATE_MISMATCH", row_number=1,
                message="找不到「名稱」/「類型」/「幣別」欄位，請使用最新範本",
            )
        ]

    out: list[ImportAccount] = []
    errors: list[ImpError] = []
    seen: set[str] = set()
    for offset, raw in enumerate(rows_2d[1:], start=2):
        if not raw or all((c or "").strip() == "" for c in raw):
            continue
        name = _cell(raw, col, "名稱")
        type_raw = _cell(raw, col, "類型")
        currency = _cell(raw, col, "幣別").upper()
        if not name:
            errors.append(ImpError(
                code="IMPORT_MISSING_NAME", row_number=offset, message="名稱為必填", field_name="name",
            ))
            continue
        if name.lower() in seen:
            errors.append(ImpError(
                code="IMPORT_DUPLICATE_ROW", row_number=offset,
                message=f"帳戶「{name}」在檔案中重複", field_name="name",
            ))
            continue
        account_type = _ACCOUNT_TYPE_ALIASES.get(type_raw.lower())
        if not account_type:
            errors.append(ImpError(
                code="IMPORT_INVALID_ACCOUNT_TYPE", row_number=offset,
                message=f"類型「{type_raw}」無法辨識，請參考範本裡的類型清單", field_name="type",
            ))
            continue
        if not currency or not (2 <= len(currency) <= 8) or not currency.isalpha():
            errors.append(ImpError(
                code="IMPORT_INVALID_CURRENCY", row_number=offset,
                message=f"幣別「{currency}」格式錯誤", field_name="currency",
            ))
            continue

        def _num(field_key: str, field_name: str) -> tuple[float | None, ImpError | None]:
            raw_val = _cell(raw, col, field_key)
            if not raw_val:
                return None, None
            try:
                return float(raw_val.replace(",", "")), None
            except ValueError:
                return None, ImpError(
                    code="IMPORT_INVALID_NUMBER", row_number=offset,
                    message=f"{field_key}「{raw_val}」不是數字", field_name=field_name,
                )

        initial_balance, err = _num("期初餘額", "initial_balance")
        if err:
            errors.append(err)
            continue
        credit_limit, err = _num("信用額度", "credit_limit")
        if err:
            errors.append(err)
            continue
        billing_day_f, err = _num("帳單日", "billing_day")
        if err:
            errors.append(err)
            continue
        payment_due_day_f, err = _num("繳款日", "payment_due_day")
        if err:
            errors.append(err)
            continue

        parent_name = _cell(raw, col, "主帳戶名稱")
        seen.add(name.lower())
        out.append(ImportAccount(
            name=name,
            type=account_type,
            currency=currency,
            initial_balance=initial_balance,
            parent_account_name=parent_name or None,
            credit_limit=credit_limit,
            billing_day=int(billing_day_f) if billing_day_f is not None else None,
            payment_due_day=int(payment_due_day_f) if payment_due_day_f is not None else None,
            source_row_number=offset,
        ))
    return out, errors
